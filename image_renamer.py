import os
import shutil
import tkinter as tk
import openpyxl
import re
from tkinter import filedialog, messagebox, scrolledtext
from pathlib import Path

class ImageRenamerApp:
    """
    /**
     * @class ImageRenamerApp
     * @description 图片重命名工具的主类，包含 GUI 界面和逻辑处理
     */
    """
    def __init__(self, root):
        """
        /**
         * @constructor
         * @param {tk.Tk} root - Tkinter 根窗口对象
         * @description 初始化界面组件
         */
        """
        self.root = root
        self.root.title("图片重命名工具")
        self.root.geometry("600x450")

        # 选中的文件夹路径
        self.target_dir = tk.StringVar()

        self._setup_ui()

    def _setup_ui(self):
        """
        /**
         * @method _setup_ui
         * @description 设置 GUI 界面布局
         */
        """
        # 路径选择区域
        path_frame = tk.Frame(self.root, pady=20)
        path_frame.pack(fill=tk.X, padx=20)

        tk.Label(path_frame, text="大文件夹路径:").pack(side=tk.LEFT)
        tk.Entry(path_frame, textvariable=self.target_dir, width=40).pack(side=tk.LEFT, padx=10)
        tk.Button(path_frame, text="选择文件夹", command=self._select_directory).pack(side=tk.LEFT)

        # 控制区域
        btn_frame = tk.Frame(self.root, pady=10)
        btn_frame.pack(fill=tk.X, padx=20)
        
        self.start_btn = tk.Button(btn_frame, text="开始执行", command=self._run_process, bg="#4CAF50", height=2)
        self.start_btn.pack(fill=tk.X)

        # 日志显示区域
        log_frame = tk.Frame(self.root, pady=10)
        log_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
        
        tk.Label(log_frame, text="运行日志:").pack(anchor=tk.W)
        self.log_area = scrolledtext.ScrolledText(log_frame, height=15)
        self.log_area.pack(fill=tk.BOTH, expand=True)

    def _select_directory(self):
        """
        /**
         * @method _select_directory
         * @description 弹出文件夹选择对话框
         */
        """
        path = filedialog.askdirectory()
        if path:
            self.target_dir.set(path)
            self._log(f"已选择目录: {path}")

    def _log(self, message):
        """
        /**
         * @method _log
         * @description 在日志区域输出信息
         * @param {string} message - 要显示的消息内容
         */
        """
        self.log_area.insert(tk.END, message + "\n")
        self.log_area.see(tk.END)
        self.root.update_idletasks()

    def _natural_sort_key(self, s):
        """
        /**
         * @method _natural_sort_key
         * @description 自然排序的键函数，确保数字按数值大小排序（如 2 排在 10 前面）
         * @param {string|Path} s - 要排序的对象
         * @returns {list} - 包含字符串和整数的列表
         */
        """
        text = str(s)
        return [int(c) if c.isdigit() else c.lower() for c in re.split(r'(\d+)', text)]

    def _run_process(self):
        """
        /**
         * @method _run_process
         * @description 执行重命名和复制的主逻辑
         */
        """
        root_path = self.target_dir.get()
        if not root_path:
            messagebox.showwarning("提示", "请先选择大文件夹路径！")
            return

        root_dir = Path(root_path)
        if not root_dir.exists():
            messagebox.showerror("错误", "所选路径不存在！")
            return

        # 创建“完成”文件夹
        output_dir = root_dir / "完成"
        try:
            output_dir.mkdir(exist_ok=True)
            self._log(f"创建/确认归档目录: {output_dir}")
        except Exception as e:
            messagebox.showerror("错误", f"无法创建完成文件夹: {str(e)}")
            return

        self.start_btn.config(state=tk.DISABLED)
        self._log("开始处理...")

        try:
            self._process_files(root_dir, output_dir)
            # 执行新增的 5 份副本功能
            self._create_extra_copies(root_dir, output_dir)
            # 执行根据 Excel 重命名的功能
            self._rename_by_excel(root_dir)
            messagebox.showinfo("成功", "所有图片处理、副本创建及 Excel 重命名完成！")
            self._log("--- 处理结束 ---")
        except Exception as e:
            self._log(f"发生错误: {str(e)}")
            messagebox.showerror("错误", f"处理过程中出现异常: {str(e)}")
        finally:
            self.start_btn.config(state=tk.NORMAL)

    def _process_files(self, root_dir, output_dir):
        """
        /**
         * @method _process_files
         * @description 核心遍历逻辑
         * @param {Path} root_dir - 大文件夹路径
         * @param {Path} output_dir - 归档文件夹路径
         */
        """
        # 使用自然排序获取所有小文件夹 (Level 1)
        small_folders = sorted(
            [f for f in root_dir.iterdir() if f.is_dir() and f.name not in ["完成", "A16", "A20", "A25", "A30", "A38"]],
            key=self._natural_sort_key
        )
        
        current_base_x = 1  # 全局起始 X
        total_count = 0

        for small_folder in small_folders:
            self._log(f"正在扫描小文件夹: {small_folder.name}")
            max_x_used_in_this_small = 0
            
            # 使用自然排序获取该小文件夹下的子文件夹 (Level 2)
            sub_folders = sorted(
                [f for f in small_folder.iterdir() if f.is_dir()],
                key=self._natural_sort_key
            )
            
            for sub_folder in sub_folders:
                y_name = sub_folder.name  # 子文件夹名即为 Y
                
                # 使用自然排序获取子文件夹下的图片文件 (Level 3)
                extensions = ('.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp', '.tiff')
                image_files = sorted(
                    [f for f in sub_folder.iterdir() if f.is_file() and f.suffix.lower() in extensions],
                    key=self._natural_sort_key
                )
                
                if not image_files:
                    continue

                self._log(f"  处理子文件夹: {y_name} (发现 {len(image_files)} 张图片)")
                
                for i, img_file in enumerate(image_files):
                    # 逻辑: X = current_base_x + 当前子文件夹内的索引
                    x_val = current_base_x + i
                    new_name = f"{x_val}-{y_name}{img_file.suffix}"
                    
                    # 复制并重命名
                    target_path = output_dir / new_name
                    
                    # 如果文件名已存在，添加序号防止覆盖（理论上按逻辑不会重复，但做个保险）
                    counter = 1
                    while target_path.exists():
                        target_path = output_dir / f"{x_val}-{y_name}_{counter}{img_file.suffix}"
                        counter += 1
                        
                    shutil.copy2(img_file, target_path)
                    
                    # 更新当前小文件夹中用到的最大 X
                    if x_val > max_x_used_in_this_small:
                        max_x_used_in_this_small = x_val
                    
                    total_count += 1
            
            # 处理完一个小文件夹后，更新下一个小文件夹的起始 base_x
            if max_x_used_in_this_small > 0:
                current_base_x = max_x_used_in_this_small + 1
            
            self._log(f"小文件夹 {small_folder.name} 处理完毕，下一个起始 X 将为 {current_base_x}")

        self._log(f"共计处理并复制了 {total_count} 张图片。")

    def _create_extra_copies(self, root_dir, source_dir):
        """
        /**
         * @method _create_extra_copies
         * @description 将“完成”文件夹内的图片排序后复制到 5 个指定的副本文件夹
         * @param {Path} root_dir - 大文件夹路径
         * @param {Path} source_dir - “完成”文件夹路径（源）
         */
        """
        copy_targets = ["A16", "A20", "A25", "A30", "A38"]
        self._log("开始创建额外的副本文件夹...")

        # 使用自然排序获取并排序“完成”文件夹内的所有文件
        files_to_copy = sorted([f for f in source_dir.iterdir() if f.is_file()], key=self._natural_sort_key)
        
        if not files_to_copy:
            self._log("“完成”文件夹中没有发现图片文件，跳过副本创建。")
            return

        for target_name in copy_targets:
            target_path = root_dir / target_name
            try:
                # 如果文件夹已存在则清空或跳过，这里选择创建（已存在则不报错）
                target_path.mkdir(exist_ok=True)
                self._log(f"  正在复制到文件夹: {target_name}")
                
                for file_path in files_to_copy:
                    shutil.copy2(file_path, target_path / file_path.name)
            except Exception as e:
                self._log(f"  复制到 {target_name} 时出错: {str(e)}")

        self._log(f"已成功创建 5 个副本文件夹: {', '.join(copy_targets)}")

    def _rename_by_excel(self, root_dir):
        """
        /**
         * @method _rename_by_excel
         * @description 根据大文件夹下的 Excel 文件内容对 5 个副本文件夹内的图片进行二次重命名
         * @param {Path} root_dir - 大文件夹路径
         */
        """
        self._log("开始根据 Excel 数据进行二次重命名...")
        
        # 1. 查找 Excel 文件（取第一个找到的 .xlsx 文件）
        excel_files = list(root_dir.glob("*.xlsx"))
        if not excel_files:
            self._log("未在大文件夹中找到 Excel (.xlsx) 文件，跳过此步骤。")
            return
        
        excel_path = excel_files[0]
        self._log(f"  使用 Excel 文件: {excel_path.name}")

        try:
            wb = openpyxl.load_workbook(excel_path, data_only=True)
        except Exception as e:
            self._log(f"  打开 Excel 出错: {str(e)}")
            return

        # 定义映射关系：子表名 -> 文件夹名
        mapping = {
            "16": "A16",
            "20": "A20",
            "25": "A25",
            "30": "A30",
            "38": "A38"
        }

        for sheet_name, folder_name in mapping.items():
            if sheet_name not in wb.sheetnames:
                self._log(f"  警告: Excel 中未找到子表 '{sheet_name}'，跳过文件夹 {folder_name}")
                continue
            
            target_folder = root_dir / folder_name
            if not target_folder.exists():
                self._log(f"  警告: 文件夹 {folder_name} 不存在，跳过")
                continue

            self._log(f"  正在根据子表 '{sheet_name}' 重命名文件夹 '{folder_name}' 中的图片...")
            
            sheet = wb[sheet_name]
            # 使用自然排序获取文件夹内的所有图片
            extensions = ('.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp', '.tiff')
            images = sorted(
                [f for f in target_folder.iterdir() if f.is_file() and f.suffix.lower() in extensions],
                key=self._natural_sort_key
            )
            
            # 读取 A 列数据（从第一行开始）
            # 假设数据在 A 列，逐行读取
            new_names = []
            for row in range(1, len(images) + 1):
                cell_value = sheet.cell(row=row, column=1).value
                if cell_value:
                    new_names.append(str(cell_value).strip())
                else:
                    new_names.append(None)

            # 执行重命名
            rename_count = 0
            for i, img_path in enumerate(images):
                if i < len(new_names) and new_names[i]:
                    new_filename = f"{new_names[i]}{img_path.suffix}"
                    new_path = target_folder / new_filename
                    
                    # 防止重名冲突
                    counter = 1
                    while new_path.exists():
                        new_path = target_folder / f"{new_names[i]}_{counter}{img_path.suffix}"
                        counter += 1
                        
                    try:
                        img_path.rename(new_path)
                        rename_count += 1
                    except Exception as e:
                        self._log(f"    重命名失败 {img_path.name}: {str(e)}")
            
            self._log(f"    文件夹 {folder_name} 完成，重命名了 {rename_count} 张图片")

        wb.close()
        self._log("Excel 二次重命名处理完毕。")

if __name__ == "__main__":
    root = tk.Tk()
    app = ImageRenamerApp(root)
    root.mainloop()
